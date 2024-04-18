import pandas as pd
import os
import sys
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

def read_capacities(file_name):
    with open(file_name, 'r') as file:
        capacities = [float(line.strip()) for line in file]
    return capacities

def write_capacities(file_name, capacities):
    with open(file_name, 'w') as file:
        for capacity in capacities:
            file.write(str(capacity) + '\n')

def assemble_battery_pack(capacities, num_series, num_parallel):
    if num_series * num_parallel > len(capacities):
        return None

    capacities.sort(reverse=True)
    battery_pack = [[] for _ in range(num_series)]
    
    for _ in range(num_parallel):
        for s in battery_pack:
            if capacities:
                s.append(capacities.pop(0))
            else:
                break
    
    return battery_pack

def calculate_voltages(num_series, chemistry):
    if chemistry.lower() == 'lion':
        cutoff_voltage_base = 2.8
        nominal_voltage_base = 3.7
        fully_charged_voltage_base = 4.2
    elif chemistry.lower() == 'lifepo4':
        cutoff_voltage_base = 2
        nominal_voltage_base = 3.2
        fully_charged_voltage_base = 3.6
    elif chemistry.lower() == 'lto':
        cutoff_voltage_base = 1.5
        nominal_voltage_base = 2.3
        fully_charged_voltage_base = 2.8    
    else:
        print("Invalid battery chemistry.")
        return None, None, None

    cutoff_voltage = round(cutoff_voltage_base * num_series, 2)
    nominal_voltage = round(nominal_voltage_base * num_series, 2)
    fully_charged_voltage = round(fully_charged_voltage_base * num_series, 2)

    return cutoff_voltage_base, cutoff_voltage, nominal_voltage_base, nominal_voltage, fully_charged_voltage_base, fully_charged_voltage

def get_integer_input(prompt):
    while True:
        try:
            value = int(input(prompt))
            return value
        except ValueError:
            print("Please enter a valid number.")

def main():
    file_name = "capacities.txt"
    if not os.path.exists(file_name):
        print("Capacities.txt file is missing. Please press enter to quit")
        input()
        sys.exit()

    capacities = read_capacities(file_name)
    
    while True:
        num_series = get_integer_input("Enter the number of series: ")
        num_parallel = get_integer_input("Enter the number of parallel cells: ")
        
        battery_chemistry = ""
        while battery_chemistry.lower() not in ['lion', 'lifepo4', 'lto']:
            battery_chemistry = input("Enter the battery chemistry (Lion/LiFePo4/LTO): ")
        
        num_packs = get_integer_input("Enter the number of packs needed: ")
        
        output_option = ""
        while output_option.lower() not in ['terminal', 'excel']:
            output_option = input("Do you want to output to terminal or save to an Excel file? (terminal/excel): ")
        
        if output_option.lower() == 'excel':
            excel_file_name = input("Enter the Excel file name: ")
            if not excel_file_name.endswith('.xlsx'):
                excel_file_name += '.xlsx'
            while os.path.exists(excel_file_name):
                overwrite = input(f"The file {excel_file_name} already exists. Overwrite existing file? (yes/no): ")
                if overwrite.lower() == 'yes':
                    break
                else:
                    excel_file_name = input("Please choose a different file name: ")
                    if not excel_file_name.endswith('.xlsx'):
                        excel_file_name += '.xlsx'
            try:
                writer = pd.ExcelWriter(excel_file_name, engine='openpyxl')
            except PermissionError:
                print("Error saving file. Please check it is not opened by another program or you have permission to save in the location.")
                return
        
        for pack_num in range(1, num_packs + 1):
            print(f"\nPack {pack_num}:")
            battery_pack = assemble_battery_pack(capacities, num_series, num_parallel)
            if battery_pack is not None:
                total_pack_capacity = 0
                series_capacities = []
                for i, series in enumerate(battery_pack, start=1):
                    total_series_capacity = sum(series)
                    total_pack_capacity += total_series_capacity
                    series_capacities.append(total_series_capacity)
                    if output_option.lower() == 'terminal':
                        print(f"Series {i}: {series}, Total Series Capacity: {total_series_capacity}")
                
                total_pack_capacity = round(total_pack_capacity / num_series, 2)
                max_series_cell_diff = round((max(series_capacities) - min(series_capacities)) / max(series_capacities) * 100, 2)
                if max_series_cell_diff > 5:
                    print("Warning: The percent difference between the highest and lowest capacity series cell is over 5%. Assemble with caution!")
                cutoff_voltage_base, cutoff_voltage, nominal_voltage_base, nominal_voltage, fully_charged_voltage_base, fully_charged_voltage = calculate_voltages(num_series, battery_chemistry)
                if output_option.lower() == 'terminal':
                    print(f"Total Pack Capacity: {total_pack_capacity}mAh, Max series cell difference: {max_series_cell_diff}%")
                    print(f"Cut Off Voltage ({cutoff_voltage_base}): {cutoff_voltage}V")
                    print(f"Nominal Voltage ({nominal_voltage_base}): {nominal_voltage}V")
                    print(f"Fully Charged Voltage ({fully_charged_voltage_base}): {fully_charged_voltage}V")
                elif output_option.lower() == 'excel':
                    
                    # Create an empty DataFrame with the correct structure
                    df = pd.DataFrame(index=[f'Series {i+1}' for i in range(num_series)] + ['Total Pack Capacity', 'Max series cell difference (%)', 'Cut Off Voltage', 'Nominal Voltage', 'Fully Charged Voltage'],
                                      columns=[f'Parallel Cell {i+1}' for i in range(num_parallel)] + ['Total Series Capacity'])
                    
                    # Fill the DataFrame with values
                    for i, series in enumerate(battery_pack, start=1):
                        for j, cell in enumerate(series, start=1):
                            df.at[f'Series {i}', f'Parallel Cell {j}'] = cell
                        df.at[f'Series {i}', 'Total Series Capacity'] = sum(series)
                    df.at['Total Pack Capacity', 'Total Series Capacity'] = total_pack_capacity
                    df.at['Max series cell difference (%)', 'Total Series Capacity'] = f"{max_series_cell_diff}%"
                    df.at['Cut Off Voltage', 'Total Series Capacity'] = f"({cutoff_voltage_base}) {cutoff_voltage}V"
                    df.at['Nominal Voltage', 'Total Series Capacity'] = f"({nominal_voltage_base}) {nominal_voltage}V"
                    df.at['Fully Charged Voltage', 'Total Series Capacity'] = f"({fully_charged_voltage_base}) {fully_charged_voltage}V"
                    # Write the DataFrame to Excel
                    df.to_excel(writer, sheet_name=f'Pack {pack_num}')
                    
                    # Open the workbook and get the sheet
                    wb = writer.book
                    ws = wb[f'Pack {pack_num}']
                    
                    # Add the sum formula for each series
                    for i in range(2, num_series + 2):
                        ws.cell(row=i, column=num_parallel + 2, value=f'=SUM(A{i}:{openpyxl.utils.get_column_letter(num_parallel + 1)}{i})')
                    
                    # Add the percent difference formula
                    col_letter = openpyxl.utils.get_column_letter(num_parallel + 2)
                    ws.cell(row=num_series + 3, column=num_parallel + 2, value=f'=ROUND((MAX({col_letter}2:{col_letter}{num_series + 1}) - MIN({col_letter}2:{col_letter}{num_series + 1})) / MAX({col_letter}2:{col_letter}{num_series + 1}) * 100, 2)')
                    
                    # Add the average total series capacity formula
                    ws.cell(row=num_series + 2, column=num_parallel + 2, value=f'=ROUND(AVERAGE({col_letter}2:{col_letter}{num_series + 1}), 2)')
                    
                print("Complete")
            else:
                print(f"Not enough cells left to create another {num_series}s{num_parallel}p pack!")
                break

        if output_option.lower() == 'excel':
            writer._save()

        remove_used = ""
        while remove_used.lower() not in ['yes', 'no']:
            remove_used = input("\nDo you want to remove the used cells from the file? (yes/no): ")
        if remove_used.lower() == 'yes':
            write_capacities(file_name, capacities)

        make_another_pack = ""
        while make_another_pack.lower() not in ['yes', 'no']:
            make_another_pack = input("\nDo you want to make another pack? (yes/no): ")
        if make_another_pack.lower() == 'no':
            break

if __name__ == "__main__":
    main()
